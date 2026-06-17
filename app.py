# pip install pandas openpyxl flask numpy
import pandas as pd
import numpy as np
import os
from flask import Flask, request, jsonify
from openpyxl import load_workbook

app = Flask(__name__)

# ===================== 基础配置 =====================
SOURCE_EXCEL = r"C:\Users\zhangjie33\Desktop\VScode\sales_data_N.xlsx"
FILTER_RULES = {
    "事业部描述": "集团食品工业事业部",
    "MKT": ["外部", "工厂自销"]
}
# 指标区分数量/金额
CORE_METRICS = [
    {"code": "sales_qty", "name": "销售数量(MT)", "label": "销量(MT)", "type": "qty"},
    {"code": "sales_revenue", "name": "销售收入", "label": "销售收入", "type": "amount"},
    {"code": "gross_profit", "name": "销售毛利(扣除运费)", "label": "销售毛利", "type": "amount"},
    {"code": "factory_cost", "name": "工厂费用", "label": "工厂费用", "type": "amount"},
    {"code": "marketing_cost", "name": "营销公司费用", "label": "营销公司费用", "type": "amount"},
    {"code": "pretax_profit", "name": "税前利润-损益（含套保）", "label": "税前利润", "type": "amount"}
]
TON_METRICS = [
    {"label": "吨均收入"},
    {"label": "吨均毛利"},
    {"label": "吨均工厂费用"},
    {"label": "吨均营销费用"},
    {"label": "吨均税前利润"}
]
MOM_DIMENSIONS = [
    {"code": "current", "label": "本月"},
    {"code": "last", "label": "上月"},
    {"code": "change", "label": "增减额"},
    {"code": "rate", "label": "变动率(%)"}
]
YOY_DIMENSIONS = [
    {"code": "current", "label": "本年累计"},
    {"code": "last", "label": "去年同期"},
    {"code": "change", "label": "增减额"},
    {"code": "rate", "label": "变动率(%)"}
]

# 全局缓存
df_cache = None

# ===================== 工具函数 =====================
def fmt_num(x, is_pct=False):
    """仅做小数保留，万元转换放到前端统一处理，后端保存原始数值"""
    if pd.isna(x) or np.isinf(x):
        val = 0.00
    else:
        val = float(x)
    if is_pct:
        return round(val * 100, 2)
    return round(val, 2)

def load_all_data():
    global df_cache
    if df_cache is not None:
        return df_cache.copy()
    fp = SOURCE_EXCEL
    if not os.path.exists(fp):
        raise Exception("源Excel文件不存在")
    wb = load_workbook(fp, read_only=True, data_only=True, keep_links=False)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header)
    col_idx = {n.strip():i for i,n in enumerate(headers) if n}
    required = [
        "事业部描述","MKT","销售组织","会计年度","期间","2017分类","所属集团重分类",
        "销售数量(MT)","销售收入","销售毛利(扣除运费)",
        "工厂税金及附加","工厂销售费用","研发费用","剩余工厂管理费用","工厂财务费用",
        "营销销售费用","营销管理费用","营销财务费用","营销附加税金","税前利润-损益（含套保）"
    ]
    miss = [f for f in required if f not in col_idx]
    if miss:
        wb.close()
        raise Exception(f"Excel缺失字段：{','.join(miss)}")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        div = str(r[col_idx["事业部描述"]]).strip() if r[col_idx["事业部描述"]] else ""
        mkt = str(r[col_idx["MKT"]]).strip() if r[col_idx["MKT"]] else ""
        if div == FILTER_RULES["事业部描述"] and mkt in FILTER_RULES["MKT"]:
            rows.append(r)
    wb.close()
    df = pd.DataFrame(rows, columns=headers)
    num_cols = [
        "销售数量(MT)","销售收入","销售毛利(扣除运费)",
        "工厂税金及附加","工厂销售费用","研发费用","剩余工厂管理费用","工厂财务费用",
        "营销销售费用","营销管理费用","营销财务费用","营销附加税金","税前利润-损益（含套保）"
    ]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df["工厂费用"] = df["工厂税金及附加"] + df["工厂销售费用"] + df["研发费用"] + df["剩余工厂管理费用"] + df["工厂财务费用"]
    df["营销公司费用"] = df["营销销售费用"] + df["营销管理费用"] + df["营销财务费用"] + df["营销附加税金"]
    df["2017分类"] = df["2017分类"].fillna("未分类产品")
    df["所属集团重分类"] = df["所属集团重分类"].fillna("未分类客户")
    df["销售组织"] = df["销售组织"].astype(str).str.strip()
    df_cache = df
    print(f"缓存加载完成，总行数{len(df)}")
    return df.copy()

# ===================== 接口 =====================
@app.route("/get_branch_list")
def get_branch():
    try:
        df = load_all_data()
        branches = sorted(df["销售组织"].unique().tolist())
        return jsonify({"branches": branches})
    except Exception as e:
        return jsonify({"error": str(e), "branches": []}),500

@app.route("/query_data", methods=["POST"])
def query_data():
    try:
        req = request.get_json()
        if not req:
            return jsonify({"error":"无参数"}),400
        target_branch = str(req.get("branch")).strip()
        target_year = int(req.get("year"))
        target_month = int(req.get("month"))
        last_year = target_year - 1
        df = load_all_data()
        df_branch = df[df["销售组织"] == target_branch].copy()
        if len(df_branch) == 0:
            return jsonify({"error":"该分公司无数据"}),400
        metric_names = [m["name"] for m in CORE_METRICS]
        amount_names = [m["name"] for m in CORE_METRICS if m["type"]=="amount"]

        # 单月/上月/累计区间
        df_curr_month = df_branch[(df_branch["会计年度"]==target_year)&(df_branch["期间"]==target_month)]
        if target_month == 1:
            df_last_month = df_branch[(df_branch["会计年度"]==target_year-1)&(df_branch["期间"]==12)]
        else:
            df_last_month = df_branch[(df_branch["会计年度"]==target_year)&(df_branch["期间"]==target_month-1)]
        df_curr_cum = df_branch[(df_branch["会计年度"]==target_year)&(df_branch["期间"]<=target_month)]
        df_last_cum = df_branch[(df_branch["会计年度"]==last_year)&(df_branch["期间"]<=target_month)]

        # 聚合函数：原始数值求和，万元转换交给前端渲染
        def agg_cat(sub_df):
            if len(sub_df)==0:
                return pd.DataFrame(columns=["2017分类"]+metric_names)
            agg = sub_df.groupby("2017分类")[metric_names].sum().reset_index()
            agg = agg.sort_values("销售数量(MT)", ascending=False).reset_index(drop=True)
            total = {"2017分类":"合计"}
            for m in metric_names:
                s = float(agg[m].sum())
                total[m] = fmt_num(s)
            agg = pd.concat([agg, pd.DataFrame([total])], ignore_index=True)
            return agg

        agg_curr_month = agg_cat(df_curr_month)
        agg_last_month = agg_cat(df_last_month)
        agg_curr_cum = agg_cat(df_curr_cum)
        agg_last_cum = agg_cat(df_last_cum)

        # 环比表格构造：后端返回原始数值，前端统一处理万元
        def build_mom(curr_df, last_df):
            data = []
            cats = curr_df["2017分类"].tolist()
            for cat in cats:
                cr = curr_df[curr_df["2017分类"]==cat].iloc[0]
                lr = last_df[last_df["2017分类"]==cat].iloc[0] if cat in last_df["2017分类"].values else pd.Series({m:0 for m in metric_names})
                row = {"分类":cat}
                for m in CORE_METRICS:
                    mn = m["name"]
                    cv = float(cr[mn])
                    lv = float(lr[mn])
                    diff = cv - lv
                    rate = diff/lv if lv !=0 else np.nan
                    row[f"{m['code']}_current"] = fmt_num(cv)
                    row[f"{m['code']}_last"] = fmt_num(lv)
                    row[f"{m['code']}_change"] = fmt_num(diff)
                    row[f"{m['code']}_rate"] = fmt_num(rate, is_pct=True)
                data.append(row)
            return data

        # 同比表格构造
        def build_yoy(curr_df, last_df):
            data = []
            cats = curr_df["2017分类"].tolist()
            for cat in cats:
                cr = curr_df[curr_df["2017分类"]==cat].iloc[0]
                lr = last_df[last_df["2017分类"]==cat].iloc[0] if cat in last_df["2017分类"].values else pd.Series({m:0 for m in metric_names})
                row = {"分类":cat}
                for m in CORE_METRICS:
                    mn = m["name"]
                    cv = float(cr[mn])
                    lv = float(lr[mn])
                    diff = cv - lv
                    rate = diff/lv if lv !=0 else np.nan
                    row[f"{m['code']}_current"] = fmt_num(cv)
                    row[f"{m['code']}_last"] = fmt_num(lv)
                    row[f"{m['code']}_change"] = fmt_num(diff)
                    row[f"{m['code']}_rate"] = fmt_num(rate, is_pct=True)
                data.append(row)
            return data

        # 吨均表格
        def build_ton(sub_df):
            data = []
            cats = sub_df["2017分类"].tolist()
            for cat in cats:
                r = sub_df[sub_df["2017分类"]==cat].iloc[0]
                qty = float(r["销售数量(MT)"])
                def ton(v):
                    return fmt_num(v/qty) if qty>0 else 0.00
                data.append({
                    "分类":cat,
                    "吨均收入":ton(r["销售收入"]),
                    "吨均毛利":ton(r["销售毛利(扣除运费)"]),
                    "吨均工厂费用":ton(r["工厂费用"]),
                    "吨均营销费用":ton(r["营销公司费用"]),
                    "吨均税前利润":ton(r["税前利润-损益（含套保）"])
                })
            return data

        mom_table = build_mom(agg_curr_month, agg_last_month)
        yoy_table = build_yoy(agg_curr_cum, agg_last_cum)
        ton_mom_curr = build_ton(agg_curr_month)
        ton_mom_last = build_ton(agg_last_month)
        ton_yoy_curr = build_ton(agg_curr_cum)
        ton_yoy_last = build_ton(agg_last_cum)

        # ========== 饼图数据（累计区间2017分类占比） ==========
        pie_base = df_curr_cum.groupby("2017分类").agg({
            "销售数量(MT)":"sum",
            "销售毛利(扣除运费)":"sum"
        }).reset_index()
        pie_qty_labels = pie_base["2017分类"].tolist()
        pie_qty_data = [float(x) for x in pie_base["销售数量(MT)"].tolist()]
        pie_gross_labels = pie_base["2017分类"].tolist()
        pie_gross_data = [float(x) for x in pie_base["销售毛利(扣除运费)"].tolist()]

        # ========== TOP20客户（所属集团重分类）：按销量降序取前20 ==========
        cust_agg = df_curr_cum.groupby("所属集团重分类").agg({
            "销售数量(MT)":"sum",
            "销售收入":"sum",
            "销售毛利(扣除运费)":"sum",
            "工厂费用":"sum",
            "营销公司费用":"sum",
            "税前利润-损益（含套保）":"sum"
        }).reset_index()
        # 按销量降序
        cust_agg = cust_agg.sort_values("销售数量(MT)", ascending=False).reset_index(drop=True)
        top20 = cust_agg.head(20).copy()
        total_all = cust_agg.sum(numeric_only=True)
        top20_sum = top20.sum(numeric_only=True)
        # 其他 = 合计 - TOP20
        other_row = {
            "所属集团重分类": "其他",
            "销售数量(MT)": fmt_num(total_all["销售数量(MT)"] - top20_sum["销售数量(MT)"]),
            "销售收入": fmt_num(total_all["销售收入"] - top20_sum["销售收入"]),
            "销售毛利(扣除运费)": fmt_num(total_all["销售毛利(扣除运费)"] - top20_sum["销售毛利(扣除运费)"]),
            "工厂费用": fmt_num(total_all["工厂费用"] - top20_sum["工厂费用"]),
            "营销公司费用": fmt_num(total_all["营销公司费用"] - top20_sum["营销公司费用"]),
            "税前利润-损益（含套保）": fmt_num(total_all["税前利润-损益（含套保）"] - top20_sum["税前利润-损益（含套保）"])
        }
        # 合计行
        total_row = {
            "所属集团重分类": "合计",
            "销售数量(MT)": fmt_num(total_all["销售数量(MT)"]),
            "销售收入": fmt_num(total_all["销售收入"]),
            "销售毛利(扣除运费)": fmt_num(total_all["销售毛利(扣除运费)"]),
            "工厂费用": fmt_num(total_all["工厂费用"]),
            "营销公司费用": fmt_num(total_all["营销公司费用"]),
            "税前利润-损益（含套保）": fmt_num(total_all["税前利润-损益（含套保）"])
        }
        cust_list = []
        for _,r in top20.iterrows():
            cust_list.append({
                "所属集团重分类": r["所属集团重分类"],
                "销售数量(MT)": fmt_num(r["销售数量(MT)"]),
                "销售收入": fmt_num(r["销售收入"]),
                "销售毛利(扣除运费)": fmt_num(r["销售毛利(扣除运费)"]),
                "工厂费用": fmt_num(r["工厂费用"]),
                "营销公司费用": fmt_num(r["营销公司费用"]),
                "税前利润-损益（含套保）": fmt_num(r["税前利润-损益（含套保）"])
            })
        cust_list.append(other_row)
        cust_list.append(total_row)

        return jsonify({
            "target_branch": target_branch,
            "year": target_year,
            "month": target_month,
            "mom_table": mom_table,
            "yoy_table": yoy_table,
            "ton_mom_curr": ton_mom_curr,
            "ton_mom_last": ton_mom_last,
            "ton_yoy_curr": ton_yoy_curr,
            "ton_yoy_last": ton_yoy_last,
            "core_metrics": CORE_METRICS,
            "mom_dimensions": MOM_DIMENSIONS,
            "yoy_dimensions": YOY_DIMENSIONS,
            "pie_qty_labels": pie_qty_labels,
            "pie_qty_data": pie_qty_data,
            "pie_gross_labels": pie_gross_labels,
            "pie_gross_data": pie_gross_data,
            "customer_table": cust_list
        })
    except Exception as e:
        print("查询异常：", str(e))
        return jsonify({"error": str(e)}),500

# 首页内嵌HTML
@app.route("/")
def index():
    html = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>分公司经营数据看板</title>
<link rel="icon" href="data:,">
<!-- 饼图数据标签插件 -->
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0;font-family:"微软雅黑"}
body{padding:20px;background:#f5f7fa;min-width:1500px}
.top-bar{background:#004085;color:#fff;padding:20px;border-radius:8px;margin-bottom:20px;box-shadow:0 2px 8px rgba(0,64,133,0.15)}
.top-bar h2{font-size:22px;font-weight:600}
.query-area{background:#fff;padding:20px;border-radius:8px;margin-bottom:20px;box-shadow:0 1px 6px rgba(0,0,0,0.08)}
.query-area>div{display:inline-block;margin-right:30px;line-height:40px}
.query-area label{font-size:14px;font-weight:600;color:#333;margin-right:8px}
select,input{padding:8px 12px;font-size:14px;width:240px;border:1px solid #d9d9d9;border-radius:4px;outline:none}
select:focus,input:focus{border-color:#007bff}
button{padding:10px 28px;background:#007bff;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:14px;font-weight:600}
button:hover{background:#0056b3}
button:disabled{background:#ccc}
.block{background:#fff;padding:24px;border-radius:8px;margin-bottom:28px;box-shadow:0 1px 6px rgba(0,0,0,0.08)}
.block h3{margin-bottom:18px;color:#003366;font-size:18px;font-weight:600;border-left:4px solid #007bff;padding-left:12px}
.table-wrap{overflow-x:auto;border-radius:6px;border:1px solid #e0e0e0}
table{width:100%;border-collapse:collapse;font-size:13px}
th,td{border:1px solid #e0e0e0;padding:10px 8px;text-align:center;vertical-align:middle}
th{background:#00509e;color:#fff;font-weight:600;position:sticky;top:0}
.level1-header{background:#004085;font-size:14px}
tr:nth-child(even){background:#f8f9fa}
tr:hover{background:#e9f5ff}
.total-row{background:#ffedb7;font-weight:600}
.cat-col{text-align:left !important;padding-left:12px;min-width:180px}
.num-col{text-align:right !important;padding-right:12px;min-width:120px}
.pct-col{text-align:right !important;padding-right:12px;min-width:110px}
.loading{color:#666;font-size:16px;padding:40px;text-align:center}
.error{color:#dc3545;padding:20px;text-align:center;background:#f8d7da;border-radius:4px}
.chart-row{display:flex;gap:30px;margin-bottom:20px}
.chart-box{flex:1;height:420px;border:1px solid #eee;border-radius:6px;padding:10px}
</style>
</head>
<body>
<div class="top-bar"><h2>分公司经营数据智能看板（含产品结构饼图+TOP20客户）</h2></div>
<div class="query-area">
    <div><label>选择分公司：</label><select id="branchSel"><option>加载中...</option></select></div>
    <div><label>统计年份：</label><input type="number" id="yearInp" value="2026" min="2010" max="2030"></div>
    <div><label>统计月份：</label><input type="number" id="monthInp" value="6" min="1" max="12"></div>
    <div><button id="searchBtn">查询数据</button></div>
</div>
<div id="resultBox"></div>
<script>
const branchSel = document.getElementById("branchSel");
const yearInp = document.getElementById("yearInp");
const monthInp = document.getElementById("monthInp");
const searchBtn = document.getElementById("searchBtn");
const resultBox = document.getElementById("resultBox");

// 会计千分位格式化
function fmtAcc(num){
    return Number(num).toLocaleString('zh-CN',{minimumFractionDigits:2,maximumFractionDigits:2,useGrouping:true})
}
function fmtPct(num){return Number(num).toFixed(2)+"%"}
// 金额转万元（前端统一转换，后端原始数值求和，保证合计精准）
function toWan(num){
    return Number(num)/10000
}

window.onload = async ()=>{
    const res = await fetch("/get_branch_list");
    const data = await res.json();
    branchSel.innerHTML = "";
    if(data.error){
        alert("加载分公司失败："+data.error);
        return;
    }
    data.branches.forEach(b=>{
        const opt = document.createElement("option");
        opt.value = b;opt.textContent = b;branchSel.appendChild(opt);
    })
}

searchBtn.onclick = async ()=>{
    const br = branchSel.value;
    const y = yearInp.value;
    const m = monthInp.value;
    if(!br){alert("请选择分公司");return}
    searchBtn.disabled = true;
    resultBox.innerHTML = "<div class='loading'>计算数据中，请稍候...</div>";
    try{
        const resp = await fetch("/query_data",{
            method:"POST",
            headers:{"Content-Type":"application/json"},
            body:JSON.stringify({branch:br,year:y,month:m})
        })
        const json = await resp.json();
        if(json.error){
            resultBox.innerHTML = `<div class='error'>${json.error}</div>`;
            return;
        }
        renderAll(json);
    }catch(e){
        resultBox.innerHTML = `<div class='error'>请求异常：${e}</div>`;
    }finally{searchBtn.disabled=false}
}

function renderAll(d){
    const y = d.year, m = d.month, ly = y-1;
    const momTitle = `${y}年${m}月 vs ${m==1?(y-1)+"12月":y+"年"+(m-1)+"月"} 单月环比`;
    const yoyTitle = `${y}年1~${m}月累计 vs ${ly}年1~${m}月累计 同比`;
    const pieHtml = `
    <div class="block">
        <h3>一、产品结构占比（累计区间2017分类）</h3>
        <div class="chart-row">
            <div class="chart-box">
                <canvas id="pieQty"></canvas>
            </div>
            <div class="chart-box">
                <canvas id="pieGross"></canvas>
            </div>
        </div>
    </div>
    `;
    let html = pieHtml;
    html += `
    <div class="block">
        <h3>二、${momTitle}</h3>
        <div class="table-wrap">${buildCoreTable(d.mom_table, d.core_metrics, d.mom_dimensions)}</div>
    </div>
    <div class="block">
        <h3>三、${yoyTitle}</h3>
        <div class="table-wrap">${buildCoreTable(d.yoy_table, d.core_metrics, d.yoy_dimensions)}</div>
    </div>
    <div class="block">
        <h3>四、吨均效益 - 单月环比</h3>
        <div class="table-wrap">${buildTonTable(d.ton_mom_curr, d.ton_mom_last)}</div>
    </div>
    <div class="block">
        <h3>五、吨均效益 - 累计同比</h3>
        <div class="table-wrap">${buildTonTable(d.ton_yoy_curr, d.ton_yoy_last)}</div>
    </div>
    <div class="block">
        <h3>六、客户结构 TOP20集团+其他（累计区间，金额单位：万元）</h3>
        <div class="table-wrap">${buildCustTable(d.customer_table)}</div>
    </div>
    `;
    resultBox.innerHTML = html;
    // 渲染饼图：关闭图例、优化标签换行与大小
    setTimeout(()=>{
        const colorList = ["#007bff","#28a745","#dc3545","#ffc107","#17a2b8","#6610f2","#fd7e14","#20c997","#6c757d","#212529"];
        // 销量饼图
        new Chart(document.getElementById("pieQty"),{
            type:"pie",
            plugins: [ChartDataLabels],
            data:{
                labels:d.pie_qty_labels,
                datasets:[{
                    label:"销量(MT)",
                    data:d.pie_qty_data,
                    backgroundColor:colorList,
                    borderWidth:1
                }]
            },
            options:{
                responsive:true,
                maintainAspectRatio:false,
                layout:{padding:20},
                plugins:{
                    title:{display:true,text:"各产品分类销量占比",font:{size:16},padding:{bottom:20}},
                    legend:{display:false}, // 完全关闭图例
                    datalabels:{
                        formatter:function(value,context){
                            const total = context.dataset.data.reduce((a,b)=>a+b,0);
                            const pct = ((value/total)*100).toFixed(1);
                            // 长名称自动换行：产品名称换行，下一行展示占比
                            return context.chart.data.labels[context.dataIndex] + "\n" + pct + "%";
                        },
                        font:{size:12,weight:"500"},
                        color:"#000",
                        anchor:"middle",
                        align:"center",
                        offset:8
                    }
                }
            }
        })
        // 毛利饼图
        new Chart(document.getElementById("pieGross"),{
            type:"pie",
            plugins: [ChartDataLabels],
            data:{
                labels:d.pie_gross_labels,
                datasets:[{
                    label:"毛利",
                    data:d.pie_gross_data,
                    backgroundColor:colorList,
                    borderWidth:1
                }]
            },
            options:{
                responsive:true,
                maintainAspectRatio:false,
                layout:{padding:20},
                plugins:{
                    title:{display:true,text:"各产品分类毛利占比",font:{size:16},padding:{bottom:20}},
                    legend:{display:false}, // 完全关闭图例
                    datalabels:{
                        formatter:function(value,context){
                            const total = context.dataset.data.reduce((a,b)=>a+b,0);
                            const pct = ((value/total)*100).toFixed(1);
                            return context.chart.data.labels[context.dataIndex] + "\n" + pct + "%";
                        },
                        font:{size:12,weight:"500"},
                        color:"#000",
                        anchor:"middle",
                        align:"center",
                        offset:8
                    }
                }
            }
        })
    },100)
}

// 核心层级表格：金额前端转万元展示
function buildCoreTable(list, metrics, dims){
    if(!list||list.length==0) return "<p>无数据</p>";
    let h1 = `<tr><th rowspan='2' class='cat-col'>2017产品分类</th>`;
    metrics.forEach(m=>h1 += `<th colspan='${dims.length}' class='level1-header'>${m.label}${m.type=="amount"?"(万元)":""}</th>`);
    h1 += "</tr>";
    let h2 = "<tr>";
    metrics.forEach(()=>dims.forEach(d=>h2+=`<th>${d.label}</th>`));
    h2 += "</tr>";
    let body = "";
    list.forEach(row=>{
        const isTotal = row["分类"]=="合计";
        let tr = `<tr class='${isTotal?"total-row":""}'><td class='cat-col'>${row["分类"]}</td>`;
        metrics.forEach(m=>{
            dims.forEach(d=>{
                const val = row[`${m.code}_${d.code}`];
                let displayVal;
                if(m.type === "amount"){
                    displayVal = fmtAcc(toWan(val));
                }else{
                    displayVal = fmtAcc(val);
                }
                tr += d.code=="rate" ? `<td class='pct-col'>${fmtPct(val)}</td>` : `<td class='num-col'>${displayVal}</td>`;
            })
        })
        tr += "</tr>"; body += tr;
    })
    return `<table><thead>${h1}${h2}</thead><tbody>${body}</tbody></table>`;
}

// 吨均表格（不转万元）
function buildTonTable(curr, last){
    if(!curr||curr.length==0) return "<p>无数据</p>";
    let head = `<tr><th rowspan='2' class='cat-col'>产品分类</th>
    <th colspan='2'>吨均收入</th><th colspan='2'>吨均毛利</th><th colspan='2'>吨均工厂费用</th><th colspan='2'>吨均营销费用</th><th colspan='2'>吨均税前利润</th></tr>
    <tr><th>本期</th><th>上期</th><th>本期</th><th>上期</th><th>本期</th><th>上期</th><th>本期</th><th>上期</th><th>本期</th><th>上期</th></tr>`;
    let body = "";
    curr.forEach(r=>{
        const cat = r["分类"];
        const lr = last.find(x=>x["分类"]==cat)||{吨均收入:0,吨均毛利:0,吨均工厂费用:0,吨均营销费用:0,吨均税前利润:0};
        const isTotal = cat=="合计";
        body += `<tr class='${isTotal?"total-row":""}'>
        <td class='cat-col'>${cat}</td>
        <td class='num-col'>${fmtAcc(r["吨均收入"])}</td><td class='num-col'>${fmtAcc(lr["吨均收入"])}</td>
        <td class='num-col'>${fmtAcc(r["吨均毛利"])}</td><td class='num-col'>${fmtAcc(lr["吨均毛利"])}</td>
        <td class='num-col'>${fmtAcc(r["吨均工厂费用"])}</td><td class='num-col'>${fmtAcc(lr["吨均工厂费用"])}</td>
        <td class='num-col'>${fmtAcc(r["吨均营销费用"])}</td><td class='num-col'>${fmtAcc(lr["吨均营销费用"])}</td>
        <td class='num-col'>${fmtAcc(r["吨均税前利润"])}</td><td class='num-col'>${fmtAcc(lr["吨均税前利润"])}</td>
        </tr>`;
    })
    return `<table><thead>${head}</thead><tbody>${body}</tbody></table>`;
}

// TOP20客户表格（前端转万元）
function buildCustTable(list){
    if(!list||list.length==0) return "<p>无数据</p>";
    let head = `<tr>
    <th class='cat-col'>所属集团重分类</th>
    <th>销量(MT)</th>
    <th>销售收入(万元)</th>
    <th>销售毛利(万元)</th>
    <th>工厂费用(万元)</th>
    <th>营销费用(万元)</th>
    <th>税前利润(万元)</th>
    </tr>`;
    let body = "";
    list.forEach(r=>{
        const isTotal = r["所属集团重分类"]=="合计";
        body += `<tr class='${isTotal?"total-row":""}'>
        <td class='cat-col'>${r["所属集团重分类"]}</td>
        <td class='num-col'>${fmtAcc(r["销售数量(MT)"])}</td>
        <td class='num-col'>${fmtAcc(toWan(r["销售收入"]))}</td>
        <td class='num-col'>${fmtAcc(toWan(r["销售毛利(扣除运费)"]))}</td>
        <td class='num-col'>${fmtAcc(toWan(r["工厂费用"]))}</td>
        <td class='num-col'>${fmtAcc(toWan(r["营销公司费用"]))}</td>
        <td class='num-col'>${fmtAcc(toWan(r["税前利润-损益（含套保）"]))}</td>
        </tr>`;
    })
    return `<table><thead>${head}</thead><tbody>${body}</tbody></table>`;
}
</script>
</body>
</html>
    """
    return html

if __name__ == "__main__":
    app.run(debug=True)